import sqlite3
import numpy as np
import re
import openpyxl
import os

from core.connect import execute_query, get_db_connection
from core.iran_city import get_name_from_code

con = sqlite3.connect('database.db')
cursorObj = con.cursor()
BASE_DIR = os.path.dirname(__file__)
IRAN_PHONE_NUMBER_FILE = os.path.join(BASE_DIR, 'data', 'phone_number.xlsx')
IRAN_MOBILE_PHONE_FILE = os.path.join(BASE_DIR, 'data', 'mobile_phone_number.xlsx')

def _get_city_name(phone_number: str) -> dict:
    """
    Get city and region name from first 3 digits of phone number

    Args:
        phone_number: 11-digit number

    Returns:
        city and region name or empty string if not found
    """
    code = int(str(phone_number)[:3])

    try:
        # wb = openpyxl.load_workbook(IRAN_PHONE_NUMBER_FILE, read_only=True, data_only=True)
        # ws = wb.active

        rows = execute_query('SELECT * FROM tPhone')
        for row in rows:
            if row[3] and row[1] and int(row[1]) == code and int(row[3]) == int(str(phone_number)[3:3+len(str(row[3]))]):
                    city = get_name_from_code(str(row[4]))
                    # wb.close()
                    return {
                        'city': city['city'],
                        'section': city['section'],
                        'county': city['county'],
                        'state': city['state'],
                        'city_code': str(row[4])
                    }
            if not (row[3]) and row[1] and int(row[1]) == code:
                # wb.close()
                city = get_name_from_code(str(row[4]))
                return {
                    'city': city['city'],
                    'section': city['section'],
                    'county': city['county'],
                    'state': city['state'],
                    'city_code': str(row[4])
                }

        # wb.close()
    except Exception as e:
        raise ValueError(f"Error reading code city data file: {str(e)}")

    return ""

def landline_phone_number_filter(phone_number: str) -> dict:
    """
    Validate Iranian landline phone number and return city information

    Args:
        phone_number: phone number string (can contain spaces, dashes, etc.)

    Returns:
        Dictionary with phone number with city and region name

    Raises:
        ValueError: If code number is invalid
    """
    cleaned = ''.join(re.findall(r"[0-9]+", str(phone_number)))
    if 13 < len(cleaned) or len(cleaned) < 10:
        raise ValueError("phone number length is incorrect!")
    if len(cleaned) > 11:
        if cleaned[:2] != "98":
            raise ValueError("phone number is not for IRAN!")
        cleaned = cleaned[2:]
    if len(cleaned) == 10:
        cleaned = "0" + cleaned
    if cleaned[0] != "0":
        raise ValueError("first number should be 0")
    name = _get_city_name(str(cleaned))
    if not name:
        raise ValueError("City not found for this phone number")
    return {
        'phone_number': str(cleaned),
        'city': name['city'],
        'section': name['section'],
        'county': name['county'],
        'state': name['state'],
        'city_code': name['city_code'],
        'type': 'landline'
    }

def _get_city_name_mobile(phone_number: str) -> dict:
    """
    Get city and operator name from first 3 digits of phone number

    Args:
        phone_number: 11-digit number

    Returns:
        city and region name or empty string if not found
    """
    try:
        # wb = openpyxl.load_workbook(IRAN_MOBILE_PHONE_FILE, read_only=True, data_only=True)
        # ws = wb.active

        rows = execute_query('SELECT * FROM tMobile')
        for row in rows:
            if row[1] and int(row[1]) == int(str(phone_number)[:len(str(row[1])) + 1]):
                    # wb.close()
                    if row[2]:
                        city = get_name_from_code(str(row[2]))
                        return {
                            'city': city['city'],
                            'section': city['section'],
                            'county': city['county'],
                            'state': city['state'],
                            'city_code': str(row[2]),
                            'operator': str(row[0])
                        }
                    else:
                        return {
                            'operator': str(row[0])
                        }

        # wb.close()
    except Exception as e:
        raise ValueError(f"Error reading code city data file: {str(e)}")

    return ""

def mobile_phone_number_filter(phone_number: str) -> dict:
    """
    Validate Iranian mobile phone number and return city information

    Args:
        phone_number: phone number string (can contain spaces, dashes, etc.)

    Returns:
        Dictionary with phone number with city and operator name

    Raises:
        ValueError: If code number is invalid
    """
    cleaned = ''.join(re.findall(r"[0-9]+", str(phone_number)))
    if 12 < len(cleaned) or len(cleaned) < 10:
        raise ValueError("phone number length is incorrect!")
    if len(cleaned) > 11:
        if cleaned[:2] != "98":
            raise ValueError("phone number is not for IRAN!")
        cleaned = cleaned[2:]
    if len(cleaned) == 10:
        cleaned = "0" + cleaned
    if not bool(re.search("^(\\+98|0)?9\\d{9}$", cleaned)):
        raise ValueError("mobile phone number format is incorrwct, there should a 9 after first 0!")
    name = _get_city_name_mobile(str(cleaned))
    if not name:
        raise ValueError("City not found for this phone number")
    if not 'city' in name:
        return {
            'phone_number': str(cleaned),
            'operator': name['operator'],
            'type': 'mobile'
        }
    else: 
        return {
            'phone_number': str(cleaned),
            'operator': name['operator'],
            'city': name['city'],
            'section': name['section'],
            'county': name['county'],
            'state': name['state'],
            'city_code': name['city_code'],
            'type': 'mobile'
        }
    
def phone_number_filter(phone_number: str) -> dict:
    """
    Validate Iranian phone number (landline or mobile) and return city information

    Args:
        phone_number: phone number string (can contain spaces, dashes, etc.)

    Returns:
        Dictionary with phone number with city and operator name

    Raises:
        ValueError: If code number is invalid
    """
    cleaned = ''.join(re.findall(r"[0-9]+", str(phone_number)))
    if len(cleaned) > 11:
        if cleaned[2] == '9':
            try:
                return mobile_phone_number_filter(str(phone_number))
            except ValueError as e:
                pass
        else:
            try:
                return landline_phone_number_filter(str(phone_number))
            except ValueError as e:
                pass
    if len(cleaned) == 10:
        cleaned = "0" + cleaned
    if cleaned[1] == '9':
        try:
            return mobile_phone_number_filter(str(phone_number))
        except ValueError as e:
            pass
    else:
        try:
            return landline_phone_number_filter(str(phone_number))
        except ValueError as e:
            pass

def get_phone_number(code: str) -> list:
    """
    Get phone number for a city with city code

    Args:
        code: 8-digit

    Returns:
        phone number in that region
    """
    
    try:
        # wb = openpyxl.load_workbook(IRAN_PHONE_NUMBER_FILE, read_only=True, data_only=True)
        # ws = wb.active
        phone = []
        rows = execute_query('SELECT * FROM tPhone')
        for row in rows:
            if row[4] and int(row[4]) == int(str(code)):
                    phone.append((str(row[1]), str(row[3])))

        # wb.close()
    except Exception as e:
        raise ValueError(f"Error reading code city data file: {str(e)}")

    return phone

def get_mobie_phone_number(code: str) -> list:
    """
    Get mobile phone number for a city with city code

    Args:
        code: 8-digit

    Returns:
        mobile phone number in that region
    """
    
    try:
        # wb = openpyxl.load_workbook(IRAN_MOBILE_PHONE_FILE, read_only=True, data_only=True)
        # ws = wb.active
        phone = []
        rows = execute_query('SELECT * FROM tMobile')
        for row in rows:
            if row[2] and int(row[2]) == int(str(code)):
                    phone.append(str(row[1]))

        # wb.close()
    except Exception as e:
        raise ValueError(f"Error reading code city data file: {str(e)}")

    return phone