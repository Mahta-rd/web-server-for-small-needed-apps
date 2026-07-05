import sqlite3
import openpyxl
import os
from sqlite3 import Error



EVENTS_FILE_PATH = os.path.join(os.path.dirname(__file__), 'events.xlsx')
try:

    con = sqlite3.connect('database.db')

except Error:

    print(Error)


cursorObj = con.cursor()

cursorObj.execute("CREATE TABLE IF NOT EXISTS tEvents(calendertype TEXT NOT NULL, month integer NOT NULL, day integer NOT NULL, desc text NOT NULL)")

con.commit()

try:

    wb = openpyxl.load_workbook(EVENTS_FILE_PATH, read_only=True, data_only=True)
    ws = wb.active

    for row in ws.iter_rows(min_row=2, values_only=True):
        data = (str(row[0]), int(row[1]), int(row[2]), str(row[3]))
        cursorObj.execute("insert into tEvents values(?, ?, ?, ?)", data)
        con.commit()
    wb.close()
except Exception as e:
    raise ValueError(f"Error reading events data file: {str(e)}")


CITY_FILE_PATH = os.path.join(os.path.dirname(__file__), 'iran_city.xlsx')

cursorObj.execute("CREATE TABLE IF NOT EXISTS tIranCity(layer integer not null, code integer primary key, parentCode integer, desc text NOT NULL, label text not null)")

con.commit()

try:

    wb = openpyxl.load_workbook(CITY_FILE_PATH, read_only=True, data_only=True)
    ws = wb.active

    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[2]:
            data = (int(row[0]), int(row[1]), int(row[2]), str(row[3]), str(row[4]))
        else:
            data = (int(row[0]), int(row[1]), None, str(row[3]), str(row[4]))
        cursorObj.execute("insert into tIranCity values(?, ?, ?, ?, ?)", data)
        con.commit()
    wb.close()
except Exception as e:
    raise ValueError(f"Error reading iran_city data file: {str(e)}")

BANK_FILE_PATH = os.path.join(os.path.dirname(__file__), 'iran_bank.xlsx')

cursorObj.execute("CREATE TABLE IF NOT EXISTS tIranBank(bankName text not null, ibanId integer not null, cardId integer)")

con.commit()

try:

    wb = openpyxl.load_workbook(BANK_FILE_PATH, read_only=True, data_only=True)
    ws = wb.active

    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[2]:
            data = (str(row[0]), int(row[1]), int(row[2]))
        else:
            data = (str(row[0]), int(row[1]), None)
        cursorObj.execute("insert into tIranBank values(?, ?, ?)", data)
        con.commit()
    wb.close()
except Exception as e:
    raise ValueError(f"Error reading iran_bank data file: {str(e)}")


MOBILE_FILE_PATH = os.path.join(os.path.dirname(__file__), 'mobile_phone_number.xlsx')

cursorObj.execute("CREATE TABLE IF NOT EXISTS tMobile(name text not null, number integer not null, cityCode integer)")

con.commit()

try:

    wb = openpyxl.load_workbook(MOBILE_FILE_PATH, read_only=True, data_only=True)
    ws = wb.active

    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[2]:
            data = (str(row[0]), int(row[1]), int(row[2]))
        else:
            data = (str(row[0]), int(row[1]), None)
        cursorObj.execute("insert into tMobile values(?, ?, ?)", data)
        con.commit()
    wb.close()
except Exception as e:
    raise ValueError(f"Error reading mobile_phone_number data file: {str(e)}")

NATIONAL_FILE_PATH = os.path.join(os.path.dirname(__file__), 'national_code_city.xlsx')

cursorObj.execute("CREATE TABLE IF NOT EXISTS tNational(code integer not null, cityName text not null, stateName text not null, cityCode integer not null)")

con.commit()

try:

    wb = openpyxl.load_workbook(NATIONAL_FILE_PATH, read_only=True, data_only=True)
    ws = wb.active

    for row in ws.iter_rows(min_row=2, max_row=608, values_only=True):
        data = (int(row[0]), str(row[1]), str(row[2]), int(row[3]))
        cursorObj.execute("insert into tNational values(?, ?, ?, ?)", data)
        con.commit()
    wb.close()
except Exception as e:
    raise ValueError(f"Error reading national_code_city data file: {str(e)}")

PHONE_FILE_PATH = os.path.join(os.path.dirname(__file__), 'phone_number.xlsx')

cursorObj.execute("CREATE TABLE IF NOT EXISTS tPhone(stateName text not null, phoneCode integer not null, regionName text not null, regionCode integer, cityCode integer not null)")

con.commit()

try:

    wb = openpyxl.load_workbook(PHONE_FILE_PATH, read_only=True, data_only=True)
    ws = wb.active

    for row in ws.iter_rows(min_row=2, max_row=779, values_only=True):
        if row[3]:
            data = (str(row[0]), int(row[1]), str(row[2]), int(row[3]), int(row[4]))
        else:
            data = (str(row[0]), int(row[1]), str(row[2]), None, int(row[4]))
        cursorObj.execute("insert into tPhone values(?, ?, ?, ?, ?)", data)
        con.commit()
    wb.close()
except Exception as e:
    raise ValueError(f"Error reading phone_number data file: {str(e)}")

PLATE_FILE_PATH = os.path.join(os.path.dirname(__file__), 'plate_number.xlsx')

cursorObj.execute("CREATE TABLE IF NOT EXISTS tPlate(stateName text not null, cityName text not null, plateNumber integer not null, plateLetter text not null, cityCode integer not null)")

con.commit()

try:

    wb = openpyxl.load_workbook(PLATE_FILE_PATH, read_only=True, data_only=True)
    ws = wb.active

    for row in ws.iter_rows(min_row=2, max_row=660, values_only=True):
        data = (str(row[0]), str(row[1]), int(row[2]), str(row[3]), int(row[4]))
        cursorObj.execute("insert into tPlate values(?, ?, ?, ?, ?)", data)
        con.commit()
    wb.close()
except Exception as e:
    raise ValueError(f"Error reading plate_number data file: {str(e)}")

POSTAL_FILE_PATH = os.path.join(os.path.dirname(__file__), 'postal_code.xlsx')

cursorObj.execute("CREATE TABLE IF NOT EXISTS tPostal(stateName text not null, cityName text not null, postalCodeStart integer not null, postalCodeEnd integer not null, cityCode integer not null)")

con.commit()

try:

    wb = openpyxl.load_workbook(POSTAL_FILE_PATH, read_only=True, data_only=True)
    ws = wb.active

    for row in ws.iter_rows(min_row=2, max_row=304, values_only=True):
        data = (str(row[0]), str(row[1]), int(row[2]), int(row[3]), int(row[4]))
        cursorObj.execute("insert into tPostal values(?, ?, ?, ?, ?)", data)
        con.commit()
    wb.close()
except Exception as e:
    raise ValueError(f"Error reading postal_code data file: {str(e)}")

LETTER_FILE_PATH = os.path.join(os.path.dirname(__file__), 'letters.xlsx')

cursorObj.execute("CREATE TABLE IF NOT EXISTS tLetter(letter text not null, code integer not null)")

con.commit()

try:

    wb = openpyxl.load_workbook(LETTER_FILE_PATH, read_only=True, data_only=True)
    ws = wb.active

    for row in ws.iter_rows(min_row=2, values_only=True):
        data = (str(row[0]), int(row[1]))
        cursorObj.execute("insert into tLetter values(?, ?)", data)
        con.commit()
    wb.close()
except Exception as e:
    raise ValueError(f"Error reading letter data file: {str(e)}")