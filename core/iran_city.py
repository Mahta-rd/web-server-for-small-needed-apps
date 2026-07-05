import sqlite3
import numpy as np
import re
import openpyxl
import os

from core.connect import execute_query, get_db_connection

con = sqlite3.connect('database.db')
cursorObj = con.cursor()
BASE_DIR = os.path.dirname(__file__)
IRAN_CITY_FILE = os.path.join(BASE_DIR, 'data', 'iran_city.xlsx')

def _get_state_name(state_code: str) -> dict:
    """
    Get state name from code

    Args:
        state_code: 2-digit number

    Returns:
        state name or empty string if not found
    """

    try:
        # wb = openpyxl.load_workbook('iran_city.xlsx', read_only=True, data_only=True)
        # ws = wb.active
        rows = execute_query('SELECT * FROM tIranCity')
        for row in rows:
            if row[1] and int(row[1]) == int(str(state_code)):
                    # wb.close()
                    return {
                        'state': str(row[3])
                    }

        # wb.close()
    except Exception as e:
        raise ValueError(f"Error reading code city data file: {str(e)}")

    return ""

def _get_county_name(county_code: str) -> dict:
    """
    Get county and state name from code

    Args:
        county_code: 4-digit number

    Returns:
        county and state name or empty string if not found
    """
    try:
        # wb = openpyxl.load_workbook('iran_city.xlsx', read_only=True, data_only=True)
        # ws = wb.active
        rows = execute_query('SELECT * FROM tIranCity')
        for row in rows:
            if row[1] and int(row[1]) == int(str(county_code)):
                    # wb.close()
                    return {
                        'county': str(row[3]),
                        'state': _get_state_name(str(county_code[:2]))['state']
                    }

        # wb.close()
    except Exception as e:
        raise ValueError(f"Error reading code city data file: {str(e)}")

    return ""

def _get_section_name(section_code: str) -> dict:
    """
    Get section, county and state name from code

    Args:
        section_code: 6-digit number

    Returns:
        section, county and state name or empty string if not found
    """
    try:
        # wb = openpyxl.load_workbook('iran_city.xlsx', read_only=True, data_only=True)
        # ws = wb.active
        rows = execute_query('SELECT * FROM tIranCity')
        for row in rows:
            if row[1] and int(row[1]) == int(str(section_code)):
                    # wb.close()
                    names = _get_county_name(str(section_code[:4]))
                    return {
                        'section': str(row[3]),
                        'county': names['county'],
                        'state': names['state']
                    }

        # wb.close()
    except Exception as e:
        raise ValueError(f"Error reading code city data file: {str(e)}")

    return ""

def _get_city_name(city_code: str) -> dict:
    """
    Get city, section, county and state name from first 3 digits of phone number

    Args:
        phone_number: 8-9-digit number

    Returns:
        city, section, county and state name or empty string if not found
    """
    try:
        # wb = openpyxl.load_workbook('iran_city.xlsx', read_only=True, data_only=True)
        # ws = wb.active
        rows = execute_query('SELECT * FROM tIranCity')
        for row in rows:
            if row[1] and int(row[1]) == int(str(city_code)):
                    # wb.close()
                    names = _get_section_name(str(city_code[:6]))
                    return {
                        'city': str(row[3]),
                        'section': names['section'],
                        'county': names['county'],
                        'state': names['state']
                    }

        # wb.close()
    except Exception as e:
        raise ValueError(f"Error reading code city data file: {str(e)}")

    return ""

def get_name_from_code(code: str) -> dict:
    """
    Get city information based on given code

    Args:
        code: code string (can contain spaces, dashes, etc.)

    Returns:
        Dictionary with city, state, county and section name

    Raises:
        ValueError: If code number is invalid
    """
    cleaned = ''.join(re.findall(r"[0-9]+", str(code)))
    if 9 < len(cleaned) or len(cleaned) < 2:
        raise ValueError("code length is incorrect!")
    if len(cleaned) == 2:
        try:
            return _get_state_name(str(cleaned))
        except ValueError as e:
            pass
    if len(cleaned) == 4:
        try:
            return _get_county_name(str(cleaned))
        except ValueError as e:
            pass
    if len(cleaned) == 6:
        try:
            return _get_section_name(str(cleaned))
        except ValueError as e:
            pass
    if len(cleaned) >= 8:
        try:
            return _get_city_name(str(cleaned))
        except ValueError as e:
            pass

def _get_state_names_by_parent(parent_code: str) -> dict:
    """
    Get all state information based on given parent code

    Args:
        parent_code: 1 digit number

    Returns:
        Dictionary with city, state, county and section name

    Raises:
        ValueError: If code number is invalid
    """
    try:
        # wb = openpyxl.load_workbook('iran_city.xlsx', read_only=True, data_only=True)
        # ws = wb.active
        states = {}
        rows = execute_query('SELECT * FROM tIranCity')
        for row in rows:
            if row[2] and int(row[2]) == int(str(parent_code)):
                    # wb.close()
                    states[f'{row[3]}'] = str(row[1])
            elif states.keys():
                break

        # wb.close()
        return states
    except Exception as e:
        raise ValueError(f"Error reading code city data file: {str(e)}")


def _get_county_names_by_parent(parent_code: str) -> dict:
    """
    Get all county information based on given parent code

    Args:
        parent_code: 2 digit number

    Returns:
        Dictionary with city, county and section name

    Raises:
        ValueError: If code number is invalid
    """
    try:
        # wb = openpyxl.load_workbook('iran_city.xlsx', read_only=True, data_only=True)
        # ws = wb.active
        county = {}
        rows = execute_query('SELECT * FROM tIranCity')
        for row in rows:
            if row[2] and int(row[2]) == int(str(parent_code)):
                    # wb.close()
                    county[f'{row[3]}'] = str(row[1])
            elif county.keys():
                break

        # wb.close()
        return county
    except Exception as e:
        raise ValueError(f"Error reading code city data file: {str(e)}")


def _get_section_names_by_parent(parent_code: str) -> dict:
    """
    Get all section information based on given parent code

    Args:
        parent_code: 4 digit number

    Returns:
        Dictionary with city name

    Raises:
        ValueError: If code number is invalid
    """
    try:
        # wb = openpyxl.load_workbook('iran_city.xlsx', read_only=True, data_only=True)
        # ws = wb.active
        sections = {}
        rows = execute_query('SELECT * FROM tIranCity')
        for row in rows:
            if row[2] and int(row[2]) == int(str(parent_code)):
                    # wb.close()
                    sections[f'{row[3]}'] = str(row[1])
            elif sections.keys():
                break

        # wb.close()
        return sections
    except Exception as e:
        raise ValueError(f"Error reading code city data file: {str(e)}")


def _get_city_names_by_parent(parent_code: str) -> list:
    """
    Get all city information based on given parent code

    Args:
        parent_code: 6 digit number

    Returns:
        list of city names

    Raises:
        ValueError: If code number is invalid
    """
    try:
        # wb = openpyxl.load_workbook('iran_city.xlsx', read_only=True, data_only=True)
        # ws = wb.active
        city = {}
        rows = execute_query('SELECT * FROM tIranCity')
        for row in rows:
            if row[2] and int(row[2]) == int(str(parent_code)):
                    # wb.close()
                    city[f"{row[3]}"] = str(row[1])
            elif city.keys():
                break

        # wb.close()
        return city
    except Exception as e:
        raise ValueError(f"Error reading code city data file: {str(e)}")


def get_information(code: str) -> dict:
    """
    Get city information based on given code

    Args:
        code: code string (can contain spaces, dashes, etc.)

    Returns:
        child layers of given code

    Raises:
        ValueError: If code number is invalid
    """
    cleaned = ''.join(re.findall(r"[0-9]+", str(code)))
    if 9 < len(cleaned) or len(cleaned) < 1:
        raise ValueError("code length is incorrect!")
    if len(cleaned) >= 8:
        try:
            return _get_city_name(str(cleaned))
        except ValueError as e:
            pass
    if len(cleaned) == 6:
        try:
            return _get_city_names_by_parent(str(cleaned))
        except ValueError as e:
            pass
    if len(cleaned) == 4:
        try:
            return _get_section_names_by_parent(str(cleaned))
        except ValueError as e:
            pass
    if len(cleaned) == 2:
        try:
            return _get_county_names_by_parent(str(cleaned))
        except ValueError as e:
            pass
    try:
        return _get_state_names_by_parent(str(cleaned))
    except ValueError as e:
        pass



print(get_information("1"))