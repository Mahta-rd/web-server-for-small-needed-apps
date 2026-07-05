"""
Module for validating Iranian bank card numbers and IBANs
"""
import sqlite3
import re
import numpy as np
import openpyxl
import os
from typing import Dict, Tuple
from core.connect import execute_query, get_db_connection

# ============================================================
# Paths to data files
# ============================================================
con = sqlite3.connect('database.db')
cursorObj = con.cursor()
BASE_DIR = os.path.dirname(__file__)
IRAN_BANK_FILE = os.path.join(BASE_DIR, 'data', 'iran_bank.xlsx')
LETTERS_FILE = os.path.join(BASE_DIR, 'data', 'letters.xlsx')


# ============================================================
# Bank Card Validation
# ============================================================

def _validate_card_luhn(card_number: str) -> bool:
    """
    Validate bank card number using Luhn algorithm

    Args:
        card_number: 16-digit card number as string

    Returns:
        True if valid, False otherwise
    """
    number = np.array([int(char) for char in str(card_number)])
    odd_mask = [i % 2 == 0 for i in range(len(str(card_number)))]
    number[odd_mask] *= 2
    number[number > 9] -= 9
    total_sum = number.sum()
    return total_sum % 10 == 0


def _get_bank_name_from_bin(card_number: str) -> str:
    """
    Get bank name from first 6 digits (BIN) of card number

    Args:
        card_number: 16-digit card number

    Returns:
        Bank name or empty string if not found
    """
    bin_prefix = str(card_number)[:6]

    try:
        # wb = openpyxl.load_workbook(IRAN_BANK_FILE, read_only=True, data_only=True)
        # ws = wb.active
        rows = execute_query('SELECT * FROM tIranBank')
        for row in rows:
            if row[2] and int(row[2]) == int(bin_prefix):
                #wb.close()
                return str(row[0])

        #wb.close()
    except Exception as e:
        raise ValueError(f"Error reading bank data file: {str(e)}")

    return ""


def check_card_number(card_number: str) -> Dict:
    """
    Validate Iranian bank card number and return bank information

    Args:
        card_number: Card number string (can contain spaces, dashes, etc.)

    Returns:
        Dictionary with card number and bank name

    Raises:
        ValueError: If card number is invalid
    """
    digits = re.findall(r"[0-9]+", str(card_number))
    clean_number = ''.join(digits)

    if len(clean_number) != 16:
        raise ValueError("Bank card number must be exactly 16 digits")

    if not _validate_card_luhn(str(clean_number)):
        raise ValueError("Invalid bank card number (Luhn check failed)")

    bank_name = _get_bank_name_from_bin(str(clean_number))
    if not bank_name:
        raise ValueError("Bank not found for this card number")

    return {
        "bank_card_number": str(clean_number),
        "bank_name": str(bank_name),
        "type": "card"
    }


# ============================================================
# IBAN Validation
# ============================================================

def _get_letter_value(letter: str) -> int:
    """
    Convert letter to its numeric value for IBAN calculation

    Args:
        letter: Single uppercase letter A-Z

    Returns:
        Numeric value (10 for A, 11 for B, ..., 35 for Z)
    """
    try:
        # wb = openpyxl.load_workbook(LETTERS_FILE, read_only=True, data_only=True)
        # ws = wb.active
        rows = execute_query('SELECT * FROM tLetter')
        for row in rows:
            if row[0] and row[0] == str(letter):
                #wb.close()
                return int(row[1])

        #wb.close()
    except Exception as e:
        raise ValueError(f"Error reading letters file: {str(e)}")

    raise ValueError(f"Invalid letter in IBAN: {str(letter)}")


def _calculate_iban_control_code(check_string: str) -> int:
    """
    Calculate IBAN control code using Mod 97-10 algorithm

    Args:
        check_string: String of digits to check

    Returns:
        Control code (should be 0-97)
    """
    r1 = int(str(check_string)[0:4]) % 97

    for i in range(4, len(str(check_string)), 4):
        block = int(str(check_string)[i:i+4])
        r2 = block % 97
        r1 = (r1 * 9 + r2) % 97

    return abs(98 - r1)


def _get_bank_name_from_iban(iban: str) -> str:
    """
    Get bank name from IBAN (first 3 digits of BBAN)

    Args:
        iban: Validated IBAN string

    Returns:
        Bank name
    """
    bban = str(iban)[4:]
    bank_code = bban[:3]

    try:
        # wb = openpyxl.load_workbook(IRAN_BANK_FILE, read_only=True, data_only=True)
        # ws = wb.active
        rows = execute_query('SELECT * FROM tIranBank')
        for row in rows:
            if row[1] and int(row[1]) == int(bank_code):
                # wb.close()
                return str(row[0])

        # wb.close()
    except Exception as e:
        raise ValueError(f"Error reading bank data file: {str(e)}")

    raise ValueError("Bank not found for this IBAN")


def check_iban(iban_str: str) -> Dict:
    """
    Validate Iranian IBAN and return bank information

    Args:
        iban_str: IBAN string (can contain spaces, dashes, etc.)

    Returns:
        Dictionary with IBAN and bank name

    Raises:
        ValueError: If IBAN is invalid
    """
    chars = re.findall(r"[a-zA-Z0-9]+", str(iban_str))
    clean_iban = ''.join(chars).upper()

    if len(clean_iban) != 26:
        raise ValueError("IBAN must be exactly 26 characters")

    if not re.match(r"[A-Z]{2}[0-9]{24}", clean_iban):
        raise ValueError("Invalid IBAN format. Expected: 2 letters followed by 24 digits")

    if clean_iban[:2] != "IR":
        raise ValueError("IBAN is not for Iran (must start with 'IR')")

    letters_num = [
        _get_letter_value(str(clean_iban)[0]),
        _get_letter_value(str(clean_iban)[1])
    ]

    check_id = clean_iban[4:] + ''.join(str(num) for num in letters_num) + "00"

    if len(check_id) % 4 != 0:
        check_id = "0" * (4 - (len(check_id) % 4)) + check_id

    control_code = _calculate_iban_control_code(str(check_id))

    if control_code != int(clean_iban[2:4]):
        raise ValueError("IBAN is invalid (control code mismatch)")

    bank_name = _get_bank_name_from_iban(str(clean_iban))

    return {
        "iban": str(clean_iban),
        "bank_name": str(bank_name),
        "type": "iban"
    }


# ============================================================
# Main validation function (auto-detect input type)
# ============================================================

def detect_and_validate(input_str: str) -> Dict:
    """
    Automatically detect whether input is a bank card number or IBAN
    and validate it

    Args:
        input_str: String containing card number or IBAN

    Returns:
        Dictionary with validation results

    Raises:
        ValueError: If input is neither valid card nor valid IBAN
    """
    cleaned = re.sub(r'[^a-zA-Z0-9]', '', str(input_str)).upper()

    if cleaned.isdigit() and len(cleaned) == 16:
        try:
            return check_card_number(cleaned)
        except ValueError as e:
            pass

    try:
        return check_iban(cleaned)
    except ValueError as e:
        pass

    raise ValueError("Input is neither a valid bank card number nor a valid IBAN")


